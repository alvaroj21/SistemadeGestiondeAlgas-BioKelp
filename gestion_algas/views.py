# -*- coding: utf-8 -*-
"""
Vistas de la aplicación de gestión de algas
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import timedelta, datetime
from functools import wraps
from .models import Usuario, TipoAlga, RegistroProduccion, ControlAcceso, CapacidadProductiva, ConfiguracionReporte
from .forms import CustomLoginForm, UsuarioCreationForm, RegistroProduccionForm, CapacidadProductivaForm, ConfiguracionReporteForm, TipoAlgaForm


# ============================================================================
# SISTEMA DE PERMISOS - CONFIGURACIÓN CENTRALIZADA
# ============================================================================

# Diccionario que define qué módulos puede acceder cada rol
PERMISOS_ROL = {
    'Administrador': [
        'dashboard',
        'registro_produccion',
        'reportes',
        'usuarios',
        'capacidad_productiva',
        'configuracion_reportes',
        'tipos_alga',
        'estadisticas_avanzadas',
    ],
    'Trabajador': [
        'dashboard',
        'registro_produccion',
        'reportes',
        'reportes_basicos',
    ],
}

# Mapeo de vistas a módulos requeridos
VISTA_MODULO = {
    'dashboard': 'dashboard',
    'registro_produccion': 'registro_produccion',
    'reportes': 'reportes',
    'usuarios': 'usuarios',
    'eliminar_usuario': 'usuarios',
    'capacidad_productiva': 'capacidad_productiva',
    'editar_capacidad': 'capacidad_productiva',
    'eliminar_capacidad': 'capacidad_productiva',
    'configuracion_reportes': 'configuracion_reportes',
    'editar_configuracion': 'configuracion_reportes',
    'eliminar_configuracion': 'configuracion_reportes',
    'generar_reporte_personalizado': 'reportes',
    'api_produccion_semanal': 'reportes',
}


# ============================================================================
# FUNCIONES AUXILIARES DE PERMISOS
# ============================================================================

def tiene_permiso(usuario, modulo):
    """
    Verifica si un usuario tiene permiso para acceder a un módulo.
    
    Args:
        usuario: Instancia de Usuario
        modulo: String con el nombre del módulo (ej: 'usuarios', 'reportes')
    
    Returns:
        bool: True si tiene permiso, False en caso contrario
    
    Ejemplo:
        if tiene_permiso(request.user, 'usuarios'):
            # Permitir acceso
    """
    if not usuario:
        return False
    
    # Verificar que tenga rol asignado
    if not usuario.rol:
        return False
    
    rol = usuario.rol
    permisos = PERMISOS_ROL.get(rol, [])
    return modulo in permisos


def verificar_permiso_vista(usuario, nombre_vista):
    """
    Verifica si un usuario puede acceder a una vista específica.
    
    Args:
        usuario: Instancia de Usuario
        nombre_vista: Nombre de la función de vista
    
    Returns:
        bool: True si tiene permiso, False en caso contrario
    """
    modulo = VISTA_MODULO.get(nombre_vista)
    if not modulo:
        return False  # Vista no mapeada = acceso denegado por defecto
    
    return tiene_permiso(usuario, modulo)


def obtener_permisos_usuario(usuario):
    """
    Obtiene lista completa de permisos de un usuario.
    Útil para debugging y auditoría.
    
    Args:
        usuario: Instancia de Usuario
    
    Returns:
        list: Lista de módulos a los que tiene acceso
    """
    if not usuario:
        return []
    
    # Verificar que sea usuario del sistema
    if not usuario.rol:
        return []
    
    rol = usuario.rol
    return PERMISOS_ROL.get(rol, [])


# ============================================================================
# DECORADORES PERSONALIZADOS DE PERMISOS
# ============================================================================

def requiere_permiso(*modulos):
    """
    Decorador para verificar permisos de acceso a vistas.
    Más legible que user_passes_test.
    
    Uso:
        @requiere_permiso('usuarios')
        def usuarios(request):
            ...
        
        @requiere_permiso('reportes', 'estadisticas_avanzadas')
        def reportes_avanzados(request):
            # Usuario debe tener AL MENOS UNO de los permisos
            ...
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            # Verificar autenticación
            if not request.session.get('user_logged', False):
                messages.error(request, 'Debes iniciar sesión para acceder a esta página')
                registrar_acceso(request, 'acceso_denegado', 
                               detalles=f'No autenticado - Vista: {view_func.__name__}')
                return redirect('login')
            
            # Obtener usuario de la sesión
            try:
                user = Usuario.objects.get(id=request.session.get('user_id'))
            except Usuario.DoesNotExist:
                request.session.flush()
                messages.error(request, 'Sesión inválida. Por favor, inicia sesión nuevamente.')
                return redirect('login')
            
            # Verificar permisos
            tiene_acceso = any(tiene_permiso(user, mod) for mod in modulos)
            
            if not tiene_acceso:
                messages.error(request, 'No tienes permisos para acceder a esta sección')
                registrar_acceso(request, 'acceso_denegado', user,
                               detalles=f'Permiso denegado - Vista: {view_func.__name__} - Rol: {user.rol}')
                return redirect('dashboard')
            
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def solo_admin(view_func):
    """
    Decorador simplificado para vistas exclusivas de administrador.
    Equivalente a @requiere_permiso('usuarios') pero más semántico.
    
    Uso:
        @solo_admin
        def usuarios(request):
            ...
    """
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.session.get('user_logged', False):
            messages.error(request, 'Debes iniciar sesión')
            return redirect('login')
        
        try:
            user = Usuario.objects.get(id=request.session.get('user_id'))
        except Usuario.DoesNotExist:
            request.session.flush()
            messages.error(request, 'Sesión inválida. Por favor, inicia sesión nuevamente.')
            return redirect('login')
        
        # Verificar que sea admin del sistema
        if user.rol != 'Administrador':
            messages.error(request, 'Esta función es exclusiva para administradores')
            registrar_acceso(request, 'acceso_denegado', user,
                           detalles=f'Acceso admin requerido - Vista: {view_func.__name__}')
            return redirect('dashboard')
        
        return view_func(request, *args, **kwargs)
    return wrapper


def permiso_lectura_escritura(modulo, requiere_escritura=False):
    """
    Decorador avanzado que diferencia entre permisos de lectura y escritura.
    
    Args:
        modulo: Nombre del módulo
        requiere_escritura: Si True, verifica permisos de escritura (admin)
    
    Uso:
        @permiso_lectura_escritura('reportes', requiere_escritura=False)
        def ver_reportes(request):
            # Admin y Trabajadores pueden ver
            ...
        
        @permiso_lectura_escritura('reportes', requiere_escritura=True)
        def eliminar_reporte(request):
            # Solo Admin puede eliminar
            ...
    """
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.session.get('user_logged', False):
                messages.error(request, 'Debes iniciar sesión')
                return redirect('login')
            
            try:
                user = Usuario.objects.get(id=request.session.get('user_id'))
            except Usuario.DoesNotExist:
                request.session.flush()
                messages.error(request, 'Sesión inválida. Por favor, inicia sesión nuevamente.')
                return redirect('login')
            
            # Verificar permiso de lectura
            if not tiene_permiso(user, modulo):
                messages.error(request, 'No tienes permisos para acceder a esta sección')
                registrar_acceso(request, 'acceso_denegado', user,
                               detalles=f'Sin permiso de lectura - Vista: {view_func.__name__}')
                return redirect('dashboard')
            
            # Si requiere escritura, verificar que sea admin
            if requiere_escritura:
                if user.rol != 'Administrador':
                    messages.error(request, 'No tienes permisos para modificar esta información')
                    registrar_acceso(request, 'acceso_denegado', user,
                                   detalles=f'Sin permiso de escritura - Vista: {view_func.__name__}')
                    return redirect('dashboard')
            
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def get_client_ip(request):
    """Obtener IP del cliente"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def registrar_acceso(request, tipo, usuario=None, detalles=None):
    """Registrar acceso en el sistema"""
    ControlAcceso.objects.create(
        usuario=usuario,
        ip_origen=get_client_ip(request),
        tipo_acceso=tipo,
        detalles=detalles
    )


def login_view(request):
    """Vista de inicio de sesión"""
    if request.session.get('user_logged', False):
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = CustomLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            
            try:
                user = Usuario.objects.get(username=username)
                if user.password == password:
                    # Iniciar sesión guardando en session
                    request.session['user_logged'] = True
                    request.session['user_id'] = user.id
                    request.session['username'] = user.username
                    request.session['rol'] = user.rol
                    request.session['email'] = user.email
                    
                    registrar_acceso(request, 'login_exitoso', user)
                    messages.success(request, f'¡Bienvenido {user.username}!')
                    return redirect('dashboard')
                else:
                    registrar_acceso(request, 'login_fallido', detalles=username)
                    messages.error(request, 'Usuario o contraseña incorrectos')
            except Usuario.DoesNotExist:
                registrar_acceso(request, 'login_fallido', detalles=username)
                messages.error(request, 'Usuario o contraseña incorrectos')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario')
    else:
        form = CustomLoginForm()
    
    return render(request, 'gestion_algas/login.html', {'form': form})


def logout_view(request):
    """Vista de cierre de sesión"""
    if request.session.get('user_logged', False):
        try:
            user_id = request.session.get('user_id')
            if user_id:
                user = Usuario.objects.get(id=user_id)
                registrar_acceso(request, 'logout', user)
        except Usuario.DoesNotExist:
            pass
    
    request.session.flush()
    messages.info(request, 'Sesión cerrada correctamente')
    return redirect('login')


@requiere_permiso('dashboard')
def dashboard(request):
    """Dashboard principal con estadísticas"""
    user = Usuario.objects.get(id=request.session.get('user_id'))
    
    # Estadísticas según el rol
    if user.rol == 'Administrador':
        # Registros del mes actual
        inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        total_registros = RegistroProduccion.objects.filter(
            fecha_registro__gte=inicio_mes
        ).count()
        
        # Producción de la última semana
        produccion_semanal = RegistroProduccion.objects.filter(
            fecha_registro__gte=timezone.now() - timedelta(days=7)
        ).aggregate(total=Sum('cantidad_cosechada'))['total'] or 0
        
        # Producción del mes actual
        produccion_total = RegistroProduccion.objects.filter(
            fecha_registro__gte=inicio_mes
        ).aggregate(total=Sum('cantidad_cosechada'))['total'] or 0
        
        ultimos_registros = RegistroProduccion.objects.select_related(
            'usuario', 'tipo_alga'
        ).order_by('-fecha_registro')[:10]
        
        # Producción por semana (últimas 4 semanas)
        produccion_por_semana = []
        etiquetas_semanas = []
        for i in range(4):
            inicio = timezone.now() - timedelta(weeks=i+1)
            fin = timezone.now() - timedelta(weeks=i)
            total_semana = RegistroProduccion.objects.filter(
                fecha_registro__gte=inicio,
                fecha_registro__lt=fin
            ).aggregate(total=Sum('cantidad_cosechada'))['total'] or 0
            produccion_por_semana.insert(0, round(float(total_semana), 2))
            # Crear etiqueta con rango de fechas
            etiqueta = f"{inicio.strftime('%d/%m')} - {fin.strftime('%d/%m')}"
            etiquetas_semanas.insert(0, etiqueta)
        
        # Producción por tipo de alga (todas las algas con producción)
        produccion_por_tipo = TipoAlga.objects.annotate(
            total=Sum('registros__cantidad_cosechada')
        ).filter(total__gt=0).order_by('-total')
        
        # Capacidad del mes actual
        capacidad_mes_actual = CapacidadProductiva.objects.filter(
            mes__year=inicio_mes.year,
            mes__month=inicio_mes.month
        ).first()
        
        if capacidad_mes_actual:
            capacidad_total = capacidad_mes_actual.capacidad_mensual_maxima
        else:
            capacidad_total = 2400  # Valor por defecto si no hay capacidad definida para el mes
        
        # Calcular porcentaje de capacidad utilizada
        if capacidad_total > 0:
            porcentaje_capacidad = round((float(produccion_total) / float(capacidad_total)) * 100, 1)
        else:
            porcentaje_capacidad = 0
            
    else:
        # Trabajador - Ver todos los datos de producción agregados (estadísticas generales)
        # pero solo sus propios registros individuales
        inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Registros totales del mes (de todos, no solo suyos)
        total_registros = RegistroProduccion.objects.filter(
            fecha_registro__gte=inicio_mes
        ).count()
        
        # Producción semanal total (de todos)
        produccion_semanal = RegistroProduccion.objects.filter(
            fecha_registro__gte=timezone.now() - timedelta(days=7)
        ).aggregate(total=Sum('cantidad_cosechada'))['total'] or 0
        
        # Producción del mes actual total (de todos)
        produccion_total = RegistroProduccion.objects.filter(
            fecha_registro__gte=inicio_mes
        ).aggregate(total=Sum('cantidad_cosechada'))['total'] or 0
        
        # Últimos registros solo del trabajador
        ultimos_registros = RegistroProduccion.objects.filter(
            usuario=user
        ).select_related('tipo_alga').order_by('-fecha_registro')[:10]
        
        # Producción por semana (últimas 4 semanas) - de todos
        produccion_por_semana = []
        etiquetas_semanas = []
        for i in range(4):
            inicio = timezone.now() - timedelta(weeks=i+1)
            fin = timezone.now() - timedelta(weeks=i)
            total_semana = RegistroProduccion.objects.filter(
                fecha_registro__gte=inicio,
                fecha_registro__lt=fin
            ).aggregate(total=Sum('cantidad_cosechada'))['total'] or 0
            produccion_por_semana.insert(0, round(float(total_semana), 2))
            # Crear etiqueta con rango de fechas
            etiqueta = f"{inicio.strftime('%d/%m')} - {fin.strftime('%d/%m')}"
            etiquetas_semanas.insert(0, etiqueta)
        
        # Producción por tipo de alga (todas las algas con producción, de todos)
        produccion_por_tipo = TipoAlga.objects.annotate(
            total=Sum('registros__cantidad_cosechada')
        ).filter(total__gt=0).order_by('-total')
        
        # Capacidad del mes actual
        capacidad_mes_actual = CapacidadProductiva.objects.filter(
            mes__year=inicio_mes.year,
            mes__month=inicio_mes.month
        ).first()
        
        if capacidad_mes_actual:
            capacidad_total = capacidad_mes_actual.capacidad_mensual_maxima
        else:
            capacidad_total = 2400  # Valor por defecto si no hay capacidad definida para el mes
        
        if capacidad_total > 0:
            porcentaje_capacidad = round((float(produccion_total) / float(capacidad_total)) * 100, 1)
        else:
            porcentaje_capacidad = 0
    
    # Obtener permisos del usuario
    permisos = obtener_permisos_usuario(user)
    
    context = {
        'total_registros': total_registros,
        'produccion_semanal': produccion_semanal,
        'produccion_total': produccion_total,
        'ultimos_registros': ultimos_registros,
        'user': user,  # Pasar objeto completo
        'username': user.username,
        'rol': user.rol,
        'permisos': permisos,
        'produccion_por_semana': produccion_por_semana,
        'etiquetas_semanas': etiquetas_semanas,
        'produccion_por_tipo': produccion_por_tipo,
        'capacidad_total': capacidad_total,
        'porcentaje_capacidad': porcentaje_capacidad,
    }
    
    return render(request, 'gestion_algas/dashboard.html', context)


@requiere_permiso('registro_produccion')
def registro_produccion(request):
    """Vista para registrar producción (Admin y Trabajador)"""
    user = Usuario.objects.get(id=request.session.get('user_id'))
    
    if request.method == 'POST':
        form = RegistroProduccionForm(request.POST)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.usuario = user
            registro.nombre_usuario = user.username
            registro.save()
            messages.success(request, 'Registro de producción guardado exitosamente!')
            return redirect('dashboard')
    else:
        form = RegistroProduccionForm()
    
    context = {
        'form': form,
        'user': user,
        'username': user.username,
        'rol': user.rol,
    }
    
    return render(request, 'gestion_algas/registro_produccion.html', context)


@solo_admin
def eliminar_registro(request, registro_id):
    """Eliminar un registro de producción (solo admin)"""
    registro = get_object_or_404(RegistroProduccion, id=registro_id)
    
    # Guardar info para el mensaje
    tipo_alga = registro.tipo_alga.nombre if registro.tipo_alga else "Desconocido"
    cantidad = registro.cantidad_cosechada
    fecha = registro.fecha_registro.strftime('%d/%m/%Y')
    
    registro.delete()
    messages.success(
        request, 
        f'Registro eliminado: {cantidad} kg de {tipo_alga} del {fecha}'
    )
    
    return redirect('dashboard')


@requiere_permiso('reportes', 'reportes_basicos')
def reportes(request):
    """Vista de reportes y estadísticas (Admin y Trabajador)"""
    user = Usuario.objects.get(id=request.session.get('user_id'))
    
    # Búsqueda
    busqueda = request.GET.get('busqueda', '')
    
    # Producción por tipo de alga
    reporte_tipos = TipoAlga.objects.annotate(
        total_cosechado=Sum('registros__cantidad_cosechada'),
        total_registros=Count('registros')
    ).filter(total_cosechado__isnull=False)
    
    if busqueda:
        reporte_tipos = reporte_tipos.filter(nombre__icontains=busqueda)
    
    # Producción por semana (últimas 8 semanas) con fechas de inicio y fin
    hace_8_semanas = timezone.now() - timedelta(weeks=8)
    registros = RegistroProduccion.objects.filter(
        fecha_registro__gte=hace_8_semanas
    ).order_by('-fecha_registro')
    
    # Agrupar por semanas de 7 días
    semanas_dict = {}
    for registro in registros:
        # Calcular el inicio de la semana (lunes)
        dias_desde_lunes = registro.fecha_registro.weekday()
        inicio_semana = (registro.fecha_registro - timedelta(days=dias_desde_lunes)).date()
        fin_semana = inicio_semana + timedelta(days=6)
        
        semana_key = inicio_semana.strftime('%Y-%m-%d')
        
        if semana_key not in semanas_dict:
            semanas_dict[semana_key] = {
                'inicio': inicio_semana,
                'fin': fin_semana,
                'total_cosechado': 0,
                'registros_count': 0
            }
        
        semanas_dict[semana_key]['total_cosechado'] += float(registro.cantidad_cosechada)
        semanas_dict[semana_key]['registros_count'] += 1
    
    # Convertir a lista ordenada
    reporte_semanas = [
        {
            'inicio': data['inicio'],
            'fin': data['fin'],
            'total_cosechado': data['total_cosechado'],
            'registros_count': data['registros_count']
        }
        for semana_key, data in sorted(semanas_dict.items(), reverse=True)
    ]
    
    # Paginacion para semanas
    paginator = Paginator(reporte_semanas, 5)
    page_number = request.GET.get('page')
    semanas_page = paginator.get_page(page_number)
    
    # Obtener permisos del usuario para mostrar/ocultar secciones
    permisos_usuario = obtener_permisos_usuario(user)
    
    context = {
        'user': user,
        'reporte_tipos': reporte_tipos,
        'reporte_semanas': semanas_page,
        'permisos_usuario': permisos_usuario,
        'busqueda': busqueda,
    }
    
    return render(request, 'gestion_algas/reportes.html', context)


@requiere_permiso('reportes')
def generar_pdf_semanal(request):
    """Genera PDF de producción semanal"""
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    from datetime import datetime
    
    # Obtener parámetros de la semana
    inicio_str = request.GET.get('inicio')
    fin_str = request.GET.get('fin')
    
    if not inicio_str or not fin_str:
        messages.error(request, 'Parámetros de fecha inválidos')
        return redirect('reportes')
    
    try:
        inicio = datetime.strptime(inicio_str, '%Y-%m-%d').date()
        fin = datetime.strptime(fin_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Formato de fecha inválido')
        return redirect('reportes')
    
    # Convertir a datetime para las consultas
    from django.utils import timezone as tz
    inicio_dt = tz.make_aware(datetime.combine(inicio, datetime.min.time()))
    fin_dt = tz.make_aware(datetime.combine(fin, datetime.max.time()))
    
    # Obtener todos los registros de la semana
    registros = RegistroProduccion.objects.filter(
        fecha_registro__gte=inicio_dt,
        fecha_registro__lte=fin_dt
    ).select_related('tipo_alga', 'usuario').order_by('fecha_registro')
    
    # Agrupar por tipo de alga
    produccion_por_tipo = {}
    total_general = 0
    
    for registro in registros:
        tipo = registro.tipo_alga.nombre
        if tipo not in produccion_por_tipo:
            produccion_por_tipo[tipo] = {
                'cantidad': 0,
                'registros': []
            }
        produccion_por_tipo[tipo]['cantidad'] += float(registro.cantidad_cosechada)
        produccion_por_tipo[tipo]['registros'].append(registro)
        total_general += float(registro.cantidad_cosechada)
    
    context = {
        'inicio': inicio,
        'fin': fin,
        'produccion_por_tipo': produccion_por_tipo,
        'total_general': total_general,
        'registros': registros,
        'fecha_generacion': timezone.now(),
    }
    
    try:
        from xhtml2pdf import pisa
        from io import BytesIO
        
        # Renderizar template
        html_string = render_to_string('gestion_algas/pdf_semanal.html', context, request=request)
        
        # Crear respuesta PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f'produccion_semanal_{inicio.strftime("%Y%m%d")}_{fin.strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Generar PDF con encoding UTF-8 usando BytesIO
        pisa_status = pisa.CreatePDF(
            BytesIO(html_string.encode('utf-8')),
            dest=response,
            encoding='utf-8'
        )
        
        if pisa_status.err:
            messages.error(request, 'Error al generar el PDF')
            return redirect('reportes')
        
        return response
    except ImportError:
        messages.error(request, 'xhtml2pdf no está instalado. Ejecuta: pip install xhtml2pdf')
        return redirect('reportes')
    except Exception as e:
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('reportes')



@solo_admin
def usuarios(request):
    """Vista de gestión de usuarios (solo admin)"""
    user = Usuario.objects.get(id=request.session.get('user_id'))
    
    if request.method == 'POST':
        form = UsuarioCreationForm(request.POST)
        if form.is_valid():
            nuevo_usuario = form.save()
            messages.success(request, f'Usuario {nuevo_usuario.username} creado exitosamente!')
            return redirect('usuarios')
    else:
        form = UsuarioCreationForm()
    
    # Búsqueda
    busqueda = request.GET.get('busqueda', '')
    lista_usuarios = Usuario.objects.all()
    
    if busqueda:
        lista_usuarios = lista_usuarios.filter(
            Q(username__icontains=busqueda) |
            Q(email__icontains=busqueda) |
            Q(rol__icontains=busqueda)
        )
    
    lista_usuarios = lista_usuarios.order_by('username')
    
    # Paginación
    paginator = Paginator(lista_usuarios, 5)
    page_number = request.GET.get('page')
    usuarios_page = paginator.get_page(page_number)
    
    context = {
        'user': user,
        'form': form,
        'usuarios': usuarios_page,
        'busqueda': busqueda,
    }
    
    return render(request, 'gestion_algas/usuarios.html', context)


@solo_admin
def eliminar_usuario(request, usuario_id):
    """Eliminar un usuario (solo admin)"""
    if request.method == 'POST':
        usuario = get_object_or_404(Usuario, id=usuario_id)
        
        # No permitir eliminar el propio usuario
        if usuario == request.user:
            messages.error(request, 'No puedes eliminar tu propio usuario')
            return redirect('usuarios')
        
        nombre = usuario.username
        usuario.delete()
        messages.success(request, f'Usuario {nombre} eliminado correctamente')
    
    return redirect('usuarios')


@login_required
def api_produccion_semanal(request):
    """API JSON para gráficos de producción semanal"""
    hace_30_dias = timezone.now() - timedelta(days=30)
    
    datos = RegistroProduccion.objects.filter(
        fecha_registro__gte=hace_30_dias
    ).extra(
        select={'fecha': 'DATE(fecha_registro)'}
    ).values('fecha').annotate(
        total=Sum('cantidad_cosechada')
    ).order_by('fecha')
    
    # Convertir a lista para JSON
    resultado = [
        {
            'fecha': str(item['fecha']),
            'total': float(item['total'])
        }
        for item in datos
    ]
    
    return JsonResponse(resultado, safe=False)


@permiso_lectura_escritura('capacidad_productiva', requiere_escritura=True)
def capacidad_productiva(request):
    """Vista de gestión de capacidad productiva (solo admin puede crear/editar)"""
    user = Usuario.objects.get(id=request.session.get('user_id'))
    
    if request.method == 'POST':
        form = CapacidadProductivaForm(request.POST)
        if form.is_valid():
            capacidad = form.save()
            messages.success(request, 'Capacidad productiva registrada exitosamente!')
            return redirect('capacidad_productiva')
    else:
        form = CapacidadProductivaForm()
    
    # Listar capacidades existentes
    capacidades = CapacidadProductiva.objects.all().order_by('-mes')
    
    context = {
        'user': user,
        'form': form,
        'capacidades': capacidades,
    }
    
    return render(request, 'gestion_algas/capacidad_productiva.html', context)


@permiso_lectura_escritura('capacidad_productiva', requiere_escritura=True)
def editar_capacidad(request, capacidad_id):
    """Editar capacidad productiva (solo admin)"""
    capacidad = get_object_or_404(CapacidadProductiva, id=capacidad_id)
    
    if request.method == 'POST':
        form = CapacidadProductivaForm(request.POST, instance=capacidad)
        if form.is_valid():
            form.save()
            messages.success(request, 'Capacidad productiva actualizada!')
            return redirect('capacidad_productiva')
    else:
        form = CapacidadProductivaForm(instance=capacidad)
    
    context = {
        'form': form,
        'capacidad': capacidad,
        'editando': True
    }
    
    return render(request, 'gestion_algas/capacidad_productiva.html', context)


@permiso_lectura_escritura('capacidad_productiva', requiere_escritura=True)
def eliminar_capacidad(request, capacidad_id):
    """Eliminar capacidad productiva (solo admin)"""
    if request.method == 'POST':
        capacidad = get_object_or_404(CapacidadProductiva, id=capacidad_id)
        mes = capacidad.mes.strftime('%m/%Y')
        capacidad.delete()
        messages.success(request, f'Capacidad del mes {mes} eliminada correctamente')
    
    return redirect('capacidad_productiva')


@permiso_lectura_escritura('configuracion_reportes', requiere_escritura=False)
def configuracion_reportes(request):
    """Vista de gestión de configuraciones de reportes (solo admin puede crear/editar)"""
    user = Usuario.objects.get(id=request.session.get('user_id'))
    
    # Solo admin puede crear/editar
    if request.method == 'POST':
        if user.rol != 'Administrador':
            messages.error(request, 'No tienes permisos para crear configuraciones')
            return redirect('configuracion_reportes')
        
        form = ConfiguracionReporteForm(request.POST)
        if form.is_valid():
            configuracion = form.save()
            messages.success(request, f'Configuración para {configuracion.empresa} creada exitosamente!')
            return redirect('configuracion_reportes')
    else:
        form = ConfiguracionReporteForm()
    
    # Búsqueda
    busqueda = request.GET.get('busqueda', '')
    configuraciones = ConfiguracionReporte.objects.all()
    
    if busqueda:
        configuraciones = configuraciones.filter(
            Q(empresa__icontains=busqueda) |
            Q(pais__icontains=busqueda) |
            Q(contacto__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )
    
    configuraciones = configuraciones.order_by('-fecha_creacion')
    
    # Paginación
    paginator = Paginator(configuraciones, 5)
    page_number = request.GET.get('page')
    configuraciones_page = paginator.get_page(page_number)
    
    # Verificar si el usuario puede editar
    puede_editar = user.rol == 'Administrador'
    
    context = {
        'user': user,
        'form': form,
        'configuraciones': configuraciones_page,
        'puede_editar': puede_editar,
        'busqueda': busqueda,
    }
    
    return render(request, 'gestion_algas/configuracion_reportes.html', context)


@permiso_lectura_escritura('configuracion_reportes', requiere_escritura=True)
def editar_configuracion(request, config_id):
    """Editar configuración de reporte (solo admin)"""
    configuracion = get_object_or_404(ConfiguracionReporte, id=config_id)
    
    # Obtener usuario desde la sesión (sistema de autenticación personalizado)
    user = Usuario.objects.get(id=request.session.get('user_id'))
    
    if request.method == 'POST':
        form = ConfiguracionReporteForm(request.POST, instance=configuracion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuración actualizada!')
            return redirect('configuracion_reportes')
    else:
        form = ConfiguracionReporteForm(instance=configuracion)
    
    puede_editar = user.rol == 'Administrador'
    
    context = {
        'form': form,
        'configuracion': configuracion,
        'editando': True,
        'puede_editar': puede_editar,
    }
    
    return render(request, 'gestion_algas/configuracion_reportes.html', context)


@permiso_lectura_escritura('configuracion_reportes', requiere_escritura=True)
def eliminar_configuracion(request, config_id):
    """Eliminar configuración de reporte (solo admin)"""
    if request.method == 'POST':
        configuracion = get_object_or_404(ConfiguracionReporte, id=config_id)
        empresa = configuracion.empresa
        configuracion.delete()
        messages.success(request, f'Configuración para {empresa} eliminada correctamente')
    
    return redirect('configuracion_reportes')


@requiere_permiso('reportes', 'configuracion_reportes')
def generar_reporte_personalizado(request, config_id):
    """Generar reporte personalizado según configuración del cliente (solo admin)"""
    configuracion = get_object_or_404(ConfiguracionReporte, id=config_id)
    
    # Determinar período de tiempo
    if configuracion.usar_fecha_personalizada and configuracion.fecha_desde and configuracion.fecha_hasta:
        fecha_desde = timezone.make_aware(datetime.combine(configuracion.fecha_desde, datetime.min.time()))
        fecha_hasta = timezone.make_aware(datetime.combine(configuracion.fecha_hasta, datetime.max.time()))
    else:
        fecha_hasta = timezone.now()
        fecha_desde = fecha_hasta - timedelta(days=30 * configuracion.periodo_historial_meses)
    
    # Obtener datos de producción con filtros
    produccion_historial = None
    produccion_por_mes = None
    if configuracion.mostrar_historial_produccion:
        query = RegistroProduccion.objects.filter(
            fecha_registro__gte=fecha_desde,
            fecha_registro__lte=fecha_hasta
        )
        
        # Filtrar por tipos de alga si están especificados
        tipos_seleccionados = configuracion.tipos_alga.all()
        if tipos_seleccionados.exists():
            query = query.filter(tipo_alga__in=tipos_seleccionados)
        
        # Filtrar por sectores si están especificados
        if configuracion.sectores_especificos:
            sectores = [s.strip() for s in configuracion.sectores_especificos.split(',')]
            query = query.filter(sector__in=sectores)
        
        produccion_historial = query.values('tipo_alga__nombre').annotate(
            total_cosechado=Sum('cantidad_cosechada'),
            total_registros=Count('id')
        ).order_by('-total_cosechado')
        
        # Datos para gráficos - usando RawSQL para compatibilidad con MySQL sin timezone
        from django.db.models import Value
        from django.db.models.functions import Concat, Substr
        
        produccion_por_mes = query.annotate(
            mes=Concat(
                Substr('fecha_registro', 1, 7),  # YYYY-MM
                Value('-01')  # Agregar día para crear fecha completa
            )
        ).values('mes').annotate(
            total=Sum('cantidad_cosechada')
        ).order_by('mes')
    
    # Factor de conversión según unidad (definir ANTES de usarlo)
    factor_conversion = 1
    if configuracion.unidad_medida == 'ton':
        factor_conversion = 0.001  # kg a toneladas
    elif configuracion.unidad_medida == 'lb':
        factor_conversion = 2.20462  # kg a libras
    
    # Obtener capacidad productiva y convertir valores
    capacidad_actual = None
    capacidad_convertida = None
    if configuracion.mostrar_capacidad_instalada or configuracion.mostrar_disponibilidad:
        # Intentar obtener la capacidad del mes del periodo final del reporte
        capacidad_actual = CapacidadProductiva.objects.filter(
            mes__year=fecha_hasta.year,
            mes__month=fecha_hasta.month
        ).first()
        
        # Si no existe capacidad para ese mes, buscar la más cercana anterior
        if not capacidad_actual:
            capacidad_actual = CapacidadProductiva.objects.filter(
                mes__lte=fecha_hasta
            ).order_by('-mes').first()
        
        if capacidad_actual:
            # Convertir valores según unidad de medida
            capacidad_convertida = {
                'capacidad_mensual_maxima': float(capacidad_actual.capacidad_mensual_maxima) * factor_conversion,
                'volumen_producido': float(capacidad_actual.volumen_producido) * factor_conversion,
                'disponibilidad_mensual': float(capacidad_actual.disponibilidad_mensual) * factor_conversion,
                'porcentaje_utilizado': capacidad_actual.porcentaje_utilizado,
                'porcentaje_disponible': capacidad_actual.porcentaje_disponible,
            }
    
    # Obtener registros detallados si incluyen observaciones
    registros_detallados = None
    if configuracion.incluir_observaciones:
        query_detalle = RegistroProduccion.objects.filter(
            fecha_registro__gte=fecha_desde,
            fecha_registro__lte=fecha_hasta
        )
        
        tipos_seleccionados = configuracion.tipos_alga.all()
        if tipos_seleccionados.exists():
            query_detalle = query_detalle.filter(tipo_alga__in=tipos_seleccionados)
        
        if configuracion.sectores_especificos:
            sectores = [s.strip() for s in configuracion.sectores_especificos.split(',')]
            query_detalle = query_detalle.filter(sector__in=sectores)
        
        registros_detallados = query_detalle.select_related('tipo_alga', 'usuario').order_by('-fecha_registro')[:50]
    
    context = {
        'configuracion': configuracion,
        'produccion_historial': produccion_historial,
        'produccion_por_mes': produccion_por_mes,
        'capacidad_actual': capacidad_actual,
        'capacidad_convertida': capacidad_convertida,
        'factor_conversion': factor_conversion,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'fecha_generacion': timezone.now(),
        'registros_detallados': registros_detallados,
    }
    
    # Renderizar según formato
    if configuracion.formato_preferido == 'pdf' or configuracion.formato_preferido == 'ambos':
        # Generar PDF con xhtml2pdf (más compatible con Windows)
        from django.template.loader import render_to_string
        from django.http import HttpResponse
        try:
            from xhtml2pdf import pisa
            from io import BytesIO
            
            # Renderizar template PDF optimizado (sin gráficos para mejor compatibilidad)
            html_string = render_to_string('gestion_algas/reporte_pdf.html', context, request=request)
            
            # Crear respuesta PDF
            response = HttpResponse(content_type='application/pdf')
            filename = 'reporte_{}_{}.pdf'.format(
                configuracion.empresa.replace(' ', '_'),
                timezone.now().strftime('%Y%m%d')
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Convertir HTML a PDF con encoding UTF-8 usando BytesIO
            pisa_status = pisa.CreatePDF(
                BytesIO(html_string.encode('utf-8')),
                dest=response,
                encoding='utf-8'
            )
            
            if pisa_status.err:
                messages.warning(request, 'Error al generar PDF. Mostrando reporte en HTML.')
                return render(request, 'gestion_algas/reporte_personalizado.html', context)
            
            return response
            
        except ImportError:
            messages.warning(request, 'xhtml2pdf no está instalado. Mostrando reporte en HTML.')
            messages.info(request, 'Para generar PDFs, instala: pip install xhtml2pdf')
            return render(request, 'gestion_algas/reporte_personalizado.html', context)
    
    elif configuracion.formato_preferido == 'excel':
        # Generar Excel
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Reporte de Producción'
            
            # Encabezado
            ws.merge_cells('A1:E1')
            header_cell = ws['A1']
            header_cell.value = f'Reporte de Producción - {configuracion.empresa}'
            header_cell.font = Font(size=16, bold=True)
            header_cell.alignment = Alignment(horizontal='center')
            
            # Información general
            row = 3
            ws[f'A{row}'] = 'País:'
            ws[f'B{row}'] = configuracion.pais
            row += 1
            ws[f'A{row}'] = 'Período:'
            periodo_texto = '{} - {}'.format(
                fecha_desde.strftime('%d/%m/%Y'),
                fecha_hasta.strftime('%d/%m/%Y')
            )
            ws[f'B{row}'] = periodo_texto
            row += 1
            ws[f'A{row}'] = 'Fecha de Generación:'
            ws[f'B{row}'] = timezone.now().strftime('%d/%m/%Y %H:%M')
            row += 2
            
            # Tabla de producción
            if produccion_historial:
                headers = ['Tipo de Alga', f'Total Cosechado ({configuracion.get_unidad_medida_display()})', 'Total Registros']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
                row += 1
                for item in produccion_historial:
                    ws.cell(row=row, column=1, value=item['tipo_alga__nombre'])
                    ws.cell(row=row, column=2, value=float(item['total_cosechado']) * factor_conversion)
                    ws.cell(row=row, column=3, value=item['total_registros'])
                    row += 1
            
            # Ajustar anchos de columna
            for col in range(1, 6):
                ws.column_dimensions[chr(64 + col)].width = 20
            
            # Crear respuesta HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = 'reporte_{}_{}.xlsx'.format(
                configuracion.empresa,
                timezone.now().strftime('%Y%m%d')
            )
            response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
            wb.save(response)
            return response
            
        except ImportError:
            messages.warning(request, 'openpyxl no está instalado. Mostrando reporte en HTML.')
            return render(request, 'gestion_algas/reporte_personalizado.html', context)
    
    else:
        return render(request, 'gestion_algas/reporte_personalizado.html', context)


@solo_admin
def tipos_alga(request):
    """Vista de gestión de tipos de alga (solo admin)"""
    user = Usuario.objects.get(id=request.session.get('user_id'))
    
    if request.method == 'POST':
        form = TipoAlgaForm(request.POST)
        if form.is_valid():
            tipo_alga = form.save()
            messages.success(request, f'Tipo de alga "{tipo_alga.nombre}" creado exitosamente!')
            return redirect('tipos_alga')
    else:
        form = TipoAlgaForm()
    
    # Búsqueda
    busqueda = request.GET.get('busqueda', '')
    lista_tipos = TipoAlga.objects.all()
    
    if busqueda:
        lista_tipos = lista_tipos.filter(
            Q(nombre__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    
    lista_tipos = lista_tipos.order_by('nombre')
    
    tipos_activos = lista_tipos.filter(activo=True).count()
    total_registros = sum(tipo.registros.count() for tipo in lista_tipos)
    
    # Paginación
    paginator = Paginator(lista_tipos, 5)
    page_number = request.GET.get('page')
    tipos_page = paginator.get_page(page_number)

    context = {
        'user': user,
        'form': form,
        'tipos_alga': tipos_page,
        'tipos_activos': tipos_activos,
        'total_registros': total_registros,
        'busqueda': busqueda,
    }

    return render(request, 'gestion_algas/tipos_alga.html', context)


@solo_admin
def editar_tipo_alga(request, tipo_id):
    """Editar un tipo de alga (solo admin)"""
    tipo_alga = get_object_or_404(TipoAlga, id=tipo_id)
    user = Usuario.objects.get(id=request.session.get('user_id'))
    
    if request.method == 'POST':
        form = TipoAlgaForm(request.POST, instance=tipo_alga)
        if form.is_valid():
            form.save()
            messages.success(request, f'Tipo de alga "{tipo_alga.nombre}" actualizado!')
            return redirect('tipos_alga')
    else:
        form = TipoAlgaForm(instance=tipo_alga)
    
    context = {
        'user': user,
        'form': form,
        'tipo_alga': tipo_alga,
        'editando': True,
        'tipos_alga': TipoAlga.objects.all().order_by('nombre'),
    }
    
    return render(request, 'gestion_algas/tipos_alga.html', context)


@solo_admin
def eliminar_tipo_alga(request, tipo_id):
    """Eliminar un tipo de alga (solo admin)"""
    if request.method == 'POST':
        tipo_alga = get_object_or_404(TipoAlga, id=tipo_id)
        
        # Guardar el nombre del tipo de alga en todos los registros antes de eliminar
        for registro in tipo_alga.registros.all():
            if not registro.nombre_tipo_alga:
                registro.nombre_tipo_alga = tipo_alga.nombre
                registro.save(update_fields=['nombre_tipo_alga'])
        
        nombre = tipo_alga.nombre
        tipo_alga.delete()
        messages.success(request, f'Tipo de alga "{nombre}" eliminado correctamente')
    
    return redirect('tipos_alga')


@solo_admin
def toggle_tipo_alga(request, tipo_id):
    """Activar/Desactivar un tipo de alga (solo admin)"""
    if request.method == 'POST':
        tipo_alga = get_object_or_404(TipoAlga, id=tipo_id)
        tipo_alga.activo = not tipo_alga.activo
        tipo_alga.save()
        
        estado = "activado" if tipo_alga.activo else "desactivado"
        messages.success(request, f'Tipo de alga "{tipo_alga.nombre}" {estado} correctamente')
    
    return redirect('tipos_alga')


# ============================================================================
# VISTA DE PERFIL/CONFIGURACIÓN DE USUARIO
# ============================================================================

@requiere_permiso('dashboard')
def perfil_usuario(request):
    """Vista para mostrar y editar el perfil del usuario logueado"""
    user = Usuario.objects.get(id=request.session.get('user_id'))
    
    if request.method == 'POST':
        # Obtener datos del formulario
        nuevo_email = request.POST.get('email', '').strip()
        nuevo_telefono = request.POST.get('telefono', '').strip()
        nueva_password = request.POST.get('password', '').strip()
        confirmar_password = request.POST.get('confirmar_password', '').strip()
        
        # Actualizar email
        if nuevo_email:
            user.email = nuevo_email
            
        # Actualizar teléfono
        if nuevo_telefono:
            user.telefono = nuevo_telefono
            
        # Actualizar contraseña si se proporcionó
        if nueva_password:
            if nueva_password != confirmar_password:
                messages.error(request, 'Las contraseñas no coinciden')
            elif len(nueva_password) < 6:
                messages.error(request, 'La contraseña debe tener al menos 6 caracteres')
            else:
                user.set_password(nueva_password)
                messages.success(request, 'Contraseña actualizada exitosamente')
        
        try:
            user.save()
            if nuevo_email or nuevo_telefono:
                messages.success(request, 'Perfil actualizado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al actualizar el perfil: {str(e)}')
        
        return redirect('perfil_usuario')
    
    context = {
        'user': user,
        'usuario': user,
        'username': user.username,
        'rol': user.rol,
    }
    
    return render(request, 'gestion_algas/perfil_usuario.html', context)

